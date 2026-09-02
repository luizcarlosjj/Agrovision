#!/usr/bin/env python3
"""
Processamento em lote para o experimento do TCC — AgroVision Grad-CAM.

Envia imagens para o backend Railway (/api/xai/gradcam) e gera um arquivo
Excel com três abas:
  Resultados — uma linha por imagem (todas as métricas brutas)
  Pareado    — uma linha por par bem+mal (diff, id_par, transformação)
  Resumo     — média ± desvio por grupo e por tipo de transformação

O descarte por confiança foi REMOVIDO (P6): todos os resultados são mantidos;
imagens abaixo de --confidence-min são marcadas na coluna 'baixa_confianca'.

Estrutura esperada em testes/:
    dataset/
        aug_metadata.csv         <- gerado por preparar_dataset.py
        bem_enquadradas/         <- JPG/PNG
        mal_enquadradas/         <- JPG/PNG
    resultados/                  <- Excel gerado aqui
    .env                         <- AGROVISION_API_KEY (nunca commitar)

Uso:
    cd testes/
    pip install -r requirements.txt
    python batch_tcc.py                         # estratégia HSV (padrão)
    python batch_tcc.py --strategy fixed        # ROI central fixa
    python batch_tcc.py --strategy hsv --strategy fixed  # não: roda uma por vez
"""

from __future__ import annotations

import argparse
import csv
import os
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path

# ─── Verificação de dependências ──────────────────────────────────────────────

missing = []
try:
    import requests
except ImportError:
    missing.append("requests")
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    missing.append("openpyxl")

if missing:
    sys.exit(
        f"Dependências ausentes: {', '.join(missing)}\n"
        f"Execute dentro de testes/: pip install -r requirements.txt"
    )

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

# ─── Constantes ───────────────────────────────────────────────────────────────

DEFAULT_API_URL  = "https://agrovision-production-bdc3.up.railway.app"
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
MAX_RETRIES      = 3
RETRY_DELAY_S    = 6

SCRIPT_DIR  = Path(__file__).parent
DATASET_DIR = SCRIPT_DIR / "dataset"
RESULTS_DIR = SCRIPT_DIR / "resultados"

COLUMNS = [
    ("arquivo",            "Arquivo"),
    ("label",              "Label"),
    ("id_par",             "ID Par"),
    ("transformacao",      "Transformação"),
    ("parametro",          "Parâmetro"),
    ("blur",               "Blur"),
    ("status",             "Status"),
    ("motivo_erro",        "Motivo Erro"),
    ("baixa_confianca",    "Baixa Confiança"),
    ("prediction",         "Predição"),
    ("confidence",         "Confiança"),
    ("al",                 "AL (Attention Leakage)"),
    ("afs",                "AFS (Focus Score)"),
    ("afs_norm",           "AFS Norm."),
    ("roi_area_frac",      "Área ROI (frac.)"),
    ("pointing_hit",       "Pointing Hit"),
    ("dist_centroide",     "Dist. Centroide"),
    ("threshold",          "Threshold"),
    ("layer_used",         "Camada Grad-CAM"),
    ("bbox_strategy",      "Estratégia BBox"),
    ("roi_fallback",       "Fallback ROI"),
    ("bbox_x",             "BBox X"),
    ("bbox_y",             "BBox Y"),
    ("bbox_w",             "BBox W"),
    ("bbox_h",             "BBox H"),
    ("image_width",        "Largura (px)"),
    ("image_height",       "Altura (px)"),
    ("processing_time_ms", "Tempo (ms)"),
    ("processado_em",      "Processado Em"),
]

STATUS_COLORS = {
    "ok":   "C6EFCE",  # verde
    "erro": "FFC7CE",  # vermelho
}

HDR_FILL  = PatternFill("solid", fgColor="1B5E20")
HDR_FONT  = Font(bold=True, color="FFFFFF")
SEC_FONT  = Font(bold=True, color="1B5E20")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")  # amarelo p/ baixa confiança


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _mime(path: Path) -> str:
    return {"png": "image/png", "webp": "image/webp"}.get(
        path.suffix.lower().lstrip("."), "image/jpeg"
    )


def _fval(d: dict, key: str, digits: int = 4):
    v = d.get(key)
    if v is None or v == "":
        return ""
    try:
        return round(float(v), digits)
    except (TypeError, ValueError):
        return v


def _avg(lst: list[dict], key: str) -> str | float:
    vals = [r[key] for r in lst if isinstance(r.get(key), (int, float))]
    return round(sum(vals) / len(vals), 4) if vals else "—"


def _std(lst: list[dict], key: str) -> str | float:
    vals = [r[key] for r in lst if isinstance(r.get(key), (int, float))]
    return round(statistics.stdev(vals), 4) if len(vals) >= 2 else "—"


def _pct_true(lst: list[dict], key: str) -> str | float:
    vals = [r[key] for r in lst if isinstance(r.get(key), bool)]
    return round(sum(vals) / len(vals) * 100, 1) if vals else "—"


# ─── Carregamento de metadados de augmentação ─────────────────────────────────

def load_aug_metadata(dataset_dir: Path) -> dict[str, dict]:
    """
    Carrega aug_metadata.csv gerado por preparar_dataset.py.
    Retorna dict keyed by mal_arquivo -> {id_par, bem_arquivo, transformacao, parametro, blur}.
    """
    meta_path = dataset_dir / "aug_metadata.csv"
    if not meta_path.exists():
        return {}
    result: dict[str, dict] = {}
    with open(meta_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            result[row["mal_arquivo"]] = row
    return result


def _id_par_from_filename(filename: str) -> str:
    """Extrai número sequencial de 'bem_003_...' ou 'mal_003_...' → '003'."""
    parts = Path(filename).stem.split("_")
    return parts[1] if len(parts) > 1 else ""


# ─── Processamento de uma imagem ──────────────────────────────────────────────

def process_image(
    image_path: Path,
    label: str,
    api_url: str,
    api_key: str,
    bbox_strategy: str,
    threshold: float,
    coverage: float,
    confidence_min: float,
) -> dict:
    empty = {k: "" for k, _ in COLUMNS
             if k not in ("arquivo", "label", "status", "motivo_erro",
                          "baixa_confianca", "processado_em",
                          "id_par", "transformacao", "parametro", "blur")}

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            with open(image_path, "rb") as fh:
                resp = requests.post(
                    f"{api_url}/api/xai/gradcam",
                    files={"image": (image_path.name, fh, _mime(image_path))},
                    data={
                        "bbox_strategy": bbox_strategy,
                        "threshold":     str(threshold),
                        "coverage":      str(coverage),
                    },
                    headers={"X-API-Key": api_key},
                    timeout=120,
                )

            if resp.status_code == 403:
                sys.exit("\n❌  API key inválida. Verifique o arquivo testes/.env")

            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:300]}")

            r    = resp.json()
            conf = float(r["confidence"])
            bbox = r.get("bbox_used", [0, 0, 0, 0])

            return {
                "arquivo":            image_path.name,
                "label":              label,
                "id_par":             "",   # preenchido depois com aug_metadata
                "transformacao":      "",
                "parametro":          "",
                "blur":               "",
                "status":             "ok",
                "motivo_erro":        "",
                "baixa_confianca":    conf < confidence_min,
                "prediction":         r.get("prediction", ""),
                "confidence":         round(conf, 4),
                "al":                 _fval(r, "al"),
                "afs":                _fval(r, "afs"),
                "afs_norm":           _fval(r, "afs_norm"),
                "roi_area_frac":      _fval(r, "roi_area_frac", 6),
                "pointing_hit":       bool(r.get("pointing_hit", False)),
                "dist_centroide":     _fval(r, "dist_centroide"),
                "threshold":          round(float(r.get("threshold", threshold)), 2),
                "layer_used":         r.get("layer_used", ""),
                "bbox_strategy":      r.get("bbox_strategy", ""),
                "roi_fallback":       bool(r.get("roi_fallback", False)),
                "bbox_x":             bbox[0] if len(bbox) > 0 else "",
                "bbox_y":             bbox[1] if len(bbox) > 1 else "",
                "bbox_w":             bbox[2] if len(bbox) > 2 else "",
                "bbox_h":             bbox[3] if len(bbox) > 3 else "",
                "image_width":        r.get("image_width", ""),
                "image_height":       r.get("image_height", ""),
                "processing_time_ms": r.get("processing_time_ms", ""),
                "processado_em":      datetime.now().isoformat(timespec="seconds"),
            }

        except (requests.Timeout, requests.ConnectionError) as exc:
            if attempt < MAX_RETRIES:
                print(f"\n     ⚠  Timeout (tentativa {attempt}/{MAX_RETRIES}). "
                      f"Aguardando {RETRY_DELAY_S}s...", flush=True)
                time.sleep(RETRY_DELAY_S)
            else:
                return {
                    "arquivo": image_path.name, "label": label,
                    "id_par": "", "transformacao": "", "parametro": "", "blur": "",
                    "status": "erro",
                    "motivo_erro": f"Timeout após {MAX_RETRIES} tentativas: {exc}",
                    "baixa_confianca": False,
                    "processado_em": datetime.now().isoformat(timespec="seconds"),
                    **empty,
                }

        except Exception as exc:
            return {
                "arquivo": image_path.name, "label": label,
                "id_par": "", "transformacao": "", "parametro": "", "blur": "",
                "status": "erro",
                "motivo_erro": str(exc)[:300],
                "baixa_confianca": False,
                "processado_em": datetime.now().isoformat(timespec="seconds"),
                **empty,
            }

    return {
        "arquivo": image_path.name, "label": label,
        "id_par": "", "transformacao": "", "parametro": "", "blur": "",
        "status": "erro", "motivo_erro": "Falha inesperada", "baixa_confianca": False,
        "processado_em": datetime.now().isoformat(), **empty,
    }


# ─── Exportação Excel ─────────────────────────────────────────────────────────

def _write_header(ws, columns: list[tuple[str, str]]) -> None:
    for col_i, (_, header) in enumerate(columns, 1):
        c = ws.cell(row=1, column=col_i, value=header)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _autowidth(ws) -> None:
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 52)


def export_excel(results: list[dict], output_path: Path, confidence_min: float) -> None:
    wb = openpyxl.Workbook()

    # ── Aba 1: Resultados ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resultados"
    _write_header(ws1, COLUMNS)

    for row_i, rec in enumerate(results, 2):
        base_fill = PatternFill("solid", fgColor=STATUS_COLORS.get(rec.get("status", ""), "FFFFFF"))
        for col_i, (key, _) in enumerate(COLUMNS, 1):
            val = rec.get(key, "")
            c = ws1.cell(row=row_i, column=col_i, value=val)
            # baixa_confianca overrides fill to yellow for ok rows
            if rec.get("baixa_confianca") and rec.get("status") == "ok":
                c.fill = WARN_FILL
            else:
                c.fill = base_fill
    _autowidth(ws1)

    # ── Aba 2: Pareado ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Pareado")

    PAR_COLS = [
        ("id_par",               "ID Par"),
        ("transformacao",        "Transformação"),
        ("parametro",            "Parâmetro"),
        ("blur",                 "Blur"),
        ("status_bem",           "Status Bem"),
        ("pred_bem",             "Predição Bem"),
        ("conf_bem",             "Confiança Bem"),
        ("al_bem",               "AL Bem"),
        ("afs_bem",              "AFS Bem"),
        ("afs_norm_bem",         "AFS Norm Bem"),
        ("roi_frac_bem",         "ROI frac Bem"),
        ("hit_bem",              "Pointing Hit Bem"),
        ("dist_bem",             "Dist. Centroide Bem"),
        ("baixa_conf_bem",       "Baixa Conf Bem"),
        ("status_mal",           "Status Mal"),
        ("pred_mal",             "Predição Mal"),
        ("conf_mal",             "Confiança Mal"),
        ("al_mal",               "AL Mal"),
        ("afs_mal",              "AFS Mal"),
        ("afs_norm_mal",         "AFS Norm Mal"),
        ("roi_frac_mal",         "ROI frac Mal"),
        ("hit_mal",              "Pointing Hit Mal"),
        ("dist_mal",             "Dist. Centroide Mal"),
        ("baixa_conf_mal",       "Baixa Conf Mal"),
        ("delta_afs",            "ΔAFS (bem−mal)"),
        ("delta_afs_norm",       "ΔAFS Norm (bem−mal)"),
        ("delta_al",             "ΔAL (mal−bem)"),
        ("delta_conf",           "ΔConf (bem−mal)"),
        ("delta_dist",           "ΔDist Centroide (mal−bem)"),
        ("par_incompleto",       "Par Incompleto"),
    ]
    _write_header(ws2, PAR_COLS)

    bem_by_par: dict[str, dict] = {}
    mal_by_par: dict[str, dict] = {}

    for rec in results:
        pid = _id_par_from_filename(rec["arquivo"])
        if not pid:
            continue
        if rec.get("label") == "bem_enquadradas":
            bem_by_par[pid] = rec
        elif rec.get("label") == "mal_enquadradas":
            mal_by_par[pid] = rec

    all_pars = sorted(set(bem_by_par) | set(mal_by_par))
    row_i = 2
    for pid in all_pars:
        bem = bem_by_par.get(pid)
        mal = mal_by_par.get(pid)

        def g(rec, key):
            return rec.get(key, "") if rec else ""

        b_ok = bem is not None and bem.get("status") == "ok"
        m_ok = mal is not None and mal.get("status") == "ok"
        incomplete = (not b_ok) or (not m_ok)

        def diff(kb, km, sign=1):
            bv = bem.get(kb) if bem else None
            mv = mal.get(km) if mal else None
            if isinstance(bv, (int, float)) and isinstance(mv, (int, float)):
                return round(sign * (bv - mv), 4)
            return ""

        par_row = {
            "id_par":          pid,
            "transformacao":   g(mal, "transformacao") or g(bem, "transformacao"),
            "parametro":       g(mal, "parametro"),
            "blur":            g(mal, "blur"),
            "status_bem":      g(bem, "status"),
            "pred_bem":        g(bem, "prediction"),
            "conf_bem":        g(bem, "confidence"),
            "al_bem":          g(bem, "al"),
            "afs_bem":         g(bem, "afs"),
            "afs_norm_bem":    g(bem, "afs_norm"),
            "roi_frac_bem":    g(bem, "roi_area_frac"),
            "hit_bem":         g(bem, "pointing_hit"),
            "dist_bem":        g(bem, "dist_centroide"),
            "baixa_conf_bem":  g(bem, "baixa_confianca"),
            "status_mal":      g(mal, "status"),
            "pred_mal":        g(mal, "prediction"),
            "conf_mal":        g(mal, "confidence"),
            "al_mal":          g(mal, "al"),
            "afs_mal":         g(mal, "afs"),
            "afs_norm_mal":    g(mal, "afs_norm"),
            "roi_frac_mal":    g(mal, "roi_area_frac"),
            "hit_mal":         g(mal, "pointing_hit"),
            "dist_mal":        g(mal, "dist_centroide"),
            "baixa_conf_mal":  g(mal, "baixa_confianca"),
            "delta_afs":       diff("afs",            "afs"),
            "delta_afs_norm":  diff("afs_norm",       "afs_norm"),
            "delta_al":        diff("al",              "al", sign=-1),
            "delta_conf":      diff("confidence",     "confidence"),
            "delta_dist":      diff("dist_centroide", "dist_centroide", sign=-1),
            "par_incompleto":  incomplete,
        }
        fill = PatternFill("solid", fgColor="FFC7CE" if incomplete else "FFFFFF")
        for col_i, (key, _) in enumerate(PAR_COLS, 1):
            c = ws2.cell(row=row_i, column=col_i, value=par_row.get(key, ""))
            c.fill = fill
        row_i += 1

    _autowidth(ws2)

    # ── Aba 3: Resumo ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Resumo")

    ok_all = [r for r in results if r.get("status") == "ok"]
    bem_ok = [r for r in ok_all if r.get("label") == "bem_enquadradas"]
    mal_ok = [r for r in ok_all if r.get("label") == "mal_enquadradas"]

    METRICS = ["confidence", "al", "afs", "afs_norm", "roi_area_frac", "dist_centroide"]

    def section(ws, start_row: int, title: str, subset: list[dict]) -> int:
        r = start_row
        c = ws.cell(row=r, column=1, value=title)
        c.font = SEC_FONT
        r += 1
        # header
        headers = ["Métrica", "N", "Média", "Desvio Padrão", "Pointing Hit (%)"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = Font(bold=True)
        r += 1
        for metric in METRICS:
            n_vals = [x for x in [rec.get(metric) for rec in subset]
                      if isinstance(x, (int, float))]
            ws.cell(row=r, column=1, value=metric)
            ws.cell(row=r, column=2, value=len(n_vals))
            ws.cell(row=r, column=3, value=_avg(subset, metric))
            ws.cell(row=r, column=4, value=_std(subset, metric))
            ws.cell(row=r, column=5, value=_pct_true(subset, "pointing_hit") if metric == "afs" else "")
            r += 1
        return r + 1  # blank line

    cur = 1
    cur = section(ws3, cur, "=== GERAL — todos status ok ===", ok_all)
    cur = section(ws3, cur, "=== BEM ENQUADRADAS ===", bem_ok)
    cur = section(ws3, cur, "=== MAL ENQUADRADAS ===", mal_ok)

    # Por tipo de transformação
    transformacoes = sorted({r.get("transformacao", "") for r in mal_ok} - {""})
    for trf in transformacoes:
        subset_trf = [r for r in mal_ok if r.get("transformacao") == trf]
        cur = section(ws3, cur, f"=== MAL — {trf} ===", subset_trf)

    # Contagens gerais
    ws3.cell(row=cur, column=1, value="=== CONTAGENS ===").font = SEC_FONT
    cur += 1
    for label, val in [
        ("Total imagens processadas",        len(results)),
        ("Status ok",                        len(ok_all)),
        ("Status erro",                      sum(1 for r in results if r.get("status") == "erro")),
        (f"Baixa confiança (< {confidence_min:.0%})",
                                             sum(1 for r in ok_all if r.get("baixa_confianca"))),
        ("Pares completos",                  sum(1 for pid in all_pars
                                                if pid in bem_by_par and pid in mal_by_par
                                                and bem_by_par[pid].get("status") == "ok"
                                                and mal_by_par[pid].get("status") == "ok")),
        ("Gerado em",                        datetime.now().isoformat(timespec="seconds")),
    ]:
        ws3.cell(row=cur, column=1, value=label)
        ws3.cell(row=cur, column=2, value=val)
        cur += 1

    ws3.column_dimensions["A"].width = 42
    for col in ["B", "C", "D", "E"]:
        ws3.column_dimensions[col].width = 18

    wb.save(output_path)


# ─── Coleta de imagens ────────────────────────────────────────────────────────

def collect_images(folder_name: str) -> list[tuple[Path, str]]:
    folder = DATASET_DIR / folder_name
    folder.mkdir(parents=True, exist_ok=True)
    imgs = sorted(p for p in folder.iterdir()
                  if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS)
    if not imgs:
        print(f"  ⚠  Pasta vazia: {folder}")
    return [(img, folder_name) for img in imgs]


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="AgroVision — Batch TCC",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--api-key",        default="",        help="AGROVISION_API_KEY (ou use .env)")
    parser.add_argument("--api-url",        default="",        help="URL do backend")
    parser.add_argument("--confidence-min", type=float, default=0.60,
                        help="Limite de confiança para marcar baixa_confianca (não descarta)")
    parser.add_argument("--threshold",      type=float, default=0.50,
                        help="Threshold do heatmap Grad-CAM (0-1)")
    parser.add_argument("--coverage",       type=float, default=0.80,
                        help="Cobertura da bounding box central (só com --strategy fixed)")
    parser.add_argument("--strategy",       default="hsv",
                        help="Estratégia de ROI: hsv (padrão) | fixed")
    parser.add_argument("--delay",          type=float, default=0.5,
                        help="Pausa entre requests (s)")
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("AGROVISION_API_KEY", "").strip()
    api_url = (args.api_url or os.environ.get("AGROVISION_API_URL", DEFAULT_API_URL)).rstrip("/")

    if not api_key:
        sys.exit(
            "\n❌  AGROVISION_API_KEY não encontrada.\n\n"
            "   Copie o arquivo .env.example para .env:\n"
            "       cp .env.example .env\n\n"
            "   Edite .env e coloque sua chave:\n"
            "       AGROVISION_API_KEY=sua_chave_aqui\n\n"
            "   (A chave está no Railway Dashboard → seu projeto → Variables)\n"
        )

    RESULTS_DIR.mkdir(exist_ok=True)
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = RESULTS_DIR / f"resultados_tcc_{timestamp}.xlsx"

    aug_meta = load_aug_metadata(DATASET_DIR)
    if aug_meta:
        print(f"  📋  aug_metadata.csv carregado: {len(aug_meta)} entradas")
    else:
        print("  ⚠  aug_metadata.csv não encontrado — colunas id_par/transformacao vazias.")
        print("      Rode: python preparar_dataset.py --etapa augmentar  (para regenerar)")

    sep = "─" * 64
    print(f"\n{sep}")
    print("  AgroVision — Processamento em Lote (TCC)")
    print(sep)
    print(f"  Backend:       {api_url}")
    print(f"  Estratégia:    bbox={args.strategy}  threshold={args.threshold}  coverage={args.coverage}")
    print(f"  Conf. mínima:  {args.confidence_min:.0%}  (abaixo → marcado, não descartado)")
    print(f"  Saída:         {output_path.relative_to(SCRIPT_DIR)}")
    print(f"{sep}\n")

    images = collect_images("bem_enquadradas") + collect_images("mal_enquadradas")

    if not images:
        sys.exit(
            "❌  Nenhuma imagem encontrada.\n"
            "   Coloque fotos JPG/PNG em:\n"
            f"     {DATASET_DIR / 'bem_enquadradas'}\n"
            f"     {DATASET_DIR / 'mal_enquadradas'}\n"
        )

    total   = len(images)
    t0      = time.perf_counter()
    results: list[dict] = []

    print(f"  {total} imagem(ns) encontrada(s). Iniciando...\n")

    try:
        for i, (img_path, label) in enumerate(images, 1):
            lbl_short = label.replace("_enquadradas", "")
            print(f"  [{i:>3}/{total}]  {lbl_short:<5}  {img_path.name:<38}", end=" ", flush=True)

            row = process_image(
                image_path=img_path,
                label=label,
                api_url=api_url,
                api_key=api_key,
                bbox_strategy=args.strategy,
                threshold=args.threshold,
                coverage=args.coverage,
                confidence_min=args.confidence_min,
            )

            # Enriquecer com metadados de augmentação (mal_enquadradas)
            if label == "mal_enquadradas":
                meta = aug_meta.get(img_path.name, {})
                # id_par = mal's own sequential number (positional pairing with bem)
                row["id_par"]        = _id_par_from_filename(img_path.name)
                row["transformacao"] = meta.get("transformacao", "")
                row["parametro"]     = meta.get("parametro", "")
                row["blur"]          = meta.get("blur", "")
            else:
                row["id_par"] = _id_par_from_filename(img_path.name)

            results.append(row)

            s = row["status"]
            lc = "⚠baixa_conf" if row.get("baixa_confianca") else ""
            if s == "ok":
                print(
                    f"✅  conf={row['confidence']:.1%}  AL={row['al']  }  "
                    f"AFS={row['afs']}  afs_norm={row['afs_norm']}  "
                    f"hit={'Y' if row['pointing_hit'] else 'N'}  "
                    f"dist={row['dist_centroide']}  {lc}"
                )
            else:
                print(f"❌  {str(row['motivo_erro'])[:80]}")

            if i < total and args.delay > 0:
                time.sleep(args.delay)

    except KeyboardInterrupt:
        print(f"\n\n  ⏹  Interrompido. Salvando {len(results)} resultado(s) parciais...")

    if not results:
        print("  Nenhum resultado para exportar.")
        return

    print(f"\n  Exportando {output_path.name}...", flush=True)
    export_excel(results, output_path, args.confidence_min)

    elapsed = time.perf_counter() - t0
    ok_n    = sum(1 for r in results if r["status"] == "ok")
    err_n   = sum(1 for r in results if r["status"] == "erro")
    lc_n    = sum(1 for r in results if r.get("baixa_confianca"))

    print(f"\n{sep}")
    print("  Concluído!")
    print(f"  ✅  OK:            {ok_n}")
    print(f"  ⚠   Baixa conf.:  {lc_n}  (conf < {args.confidence_min:.0%}, mantidos na análise)")
    print(f"  ❌  Erros:         {err_n}")
    print(f"  ⏱   Tempo total:  {elapsed:.0f}s")
    print(f"  📄  Arquivo:       {output_path}")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
