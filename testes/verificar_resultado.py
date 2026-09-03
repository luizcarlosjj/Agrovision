"""Verifica sanidade do Excel gerado: ROI area, Pareado, métricas por grupo."""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

results_dir = Path(__file__).parent / "resultados"
xlsx_files  = sorted(results_dir.glob("resultados_tcc_*.xlsx"))
if not xlsx_files:
    sys.exit("Nenhum Excel encontrado em resultados/")

latest = xlsx_files[-1]
print(f"Arquivo: {latest.name}\n")

wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)

# ── Aba Resultados ────────────────────────────────────────────────────────────
ws = wb["Resultados"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

def col(name):
    return headers.index(name)

rows = list(ws.iter_rows(min_row=2, values_only=True))

def vals(label_filter, key):
    ci = col(key)
    return [r[ci] for r in rows
            if r[col("Label")] == label_filter
            and isinstance(r[ci], (int, float))]

# ROI area frac por grupo e por transformação
print("=== roi_area_frac por grupo ===")
for label in ("bem_enquadradas", "mal_enquadradas"):
    v = vals(label, "Área ROI (frac.)")
    if v:
        print(f"  {label}: n={len(v)}  média={sum(v)/len(v):.3f}  "
              f"min={min(v):.3f}  max={max(v):.3f}")

print()
print("=== roi_area_frac das mal_enquadradas por transformação ===")
trans_col = col("Transformação")
roi_col   = col("Área ROI (frac.)")
for trf in ("zoom_out", "offset", "rotate", "partial"):
    v = [r[roi_col] for r in rows
         if r[col("Label")] == "mal_enquadradas"
         and r[trans_col] == trf
         and isinstance(r[roi_col], (int, float))]
    if v:
        print(f"  {trf:<12}: n={len(v)}  média={sum(v)/len(v):.3f}  "
              f"min={min(v):.3f}  max={max(v):.3f}")

print()
print("=== AFS e AFS Norm por grupo (status ok) ===")
status_col = col("Status")
for label in ("bem_enquadradas", "mal_enquadradas"):
    ok_rows = [r for r in rows
               if r[col("Label")] == label and r[status_col] == "ok"]
    for metric, key in [("AFS", "AFS (Focus Score)"), ("AFS Norm.", "AFS Norm."),
                         ("AL",  "AL (Attention Leakage)"),
                         ("Pointing Hit %", "Pointing Hit")]:
        ci = col(key)
        v = [r[ci] for r in ok_rows if isinstance(r[ci], (int, float, bool))]
        if not v:
            continue
        if metric == "Pointing Hit %":
            pct = sum(bool(x) for x in v) / len(v) * 100
            print(f"  {label:<22} {metric}: {pct:.1f}%")
        else:
            mean = sum(float(x) for x in v) / len(v)
            print(f"  {label:<22} {metric}: {mean:.4f}")
    print()

# ── Aba Pareado ───────────────────────────────────────────────────────────────
ws2 = wb["Pareado"]
par_rows = list(ws2.iter_rows(min_row=2, values_only=True))
n_par     = sum(1 for r in par_rows if any(c is not None for c in r))
h2        = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
inc_col   = h2.index("Par Incompleto")
n_inc     = sum(1 for r in par_rows if r[inc_col] is True)
print(f"=== Aba Pareado ===")
print(f"  Linhas de par:   {n_par}  (esperado: 50)")
print(f"  Pares incompletos: {n_inc}  (esperado: 0)")

# Verificação extra: zoom_out — ROI vs escala esperada
print()
print("=== Sanidade zoom_out: roi_area_frac vs escala^2 ===")
param_col = col("Parâmetro")
for r in rows:
    if (r[col("Label")] == "mal_enquadradas"
            and r[trans_col] == "zoom_out"
            and isinstance(r[roi_col], float)
            and r[param_col]):
        try:
            scale = float(r[param_col])
            expected = round(scale ** 2, 3)
            ratio    = round(r[roi_col] / expected, 2) if expected > 0 else "?"
            flag     = "  ⚠ GRANDE" if isinstance(ratio, float) and ratio > 3.0 else ""
            print(f"  {r[col('Arquivo')]:<45} escala={scale:.3f}  "
                  f"exp={expected:.3f}  roi={r[roi_col]:.3f}  ratio={ratio}{flag}")
        except (ValueError, TypeError):
            pass
